"""
Xuat bao cao hanh trinh TONG HOP cho TOAN BO xe tren Adsun, gop vao Google
Sheets (1 tab / thang) + 1 ban sao Excel local. Muc dich: chay tu dong MOI
GIO (qua Windows Task Scheduler) de cap nhat gan-thoi-gian-thuc.

"Ngay lam viec" (business day) tinh tu 8h sang hom nay/hom qua den THOI DIEM
CHAY (khong co dinh la 8h-8h nhu truoc) — moi lan chay se GHI DE dung phan
du lieu cua ngay lam viec hien tai trong tab thang tren Sheets, GIU NGUYEN
cac ngay khac da co san trong tab do (xem sheets_client.upsert_range_report).
Nho vay moi lan chay chi ton thoi gian quet + ghi cho 1 ngay, khong lien
quan kich thuoc ca thang.

CACH HOAT DONG:
    1. Dang nhap, lay danh sach toan bo xe (dropdown "Thiết bị").
    2. Voi moi xe: chon xe, chon khoang "2 ngày" (du de phu 8h-8h, loc chinh
       xac hon o buoc sau), bam Xem, xuat Excel rieng cho xe do (luu tam vao
       thu muc raw/).
    3. Doc tat ca file rieng, LOC lai dung khung gio cua ngay lam viec hien tai.
    4. Voi tung xe: bo cac dong khong co Vung, gop cac dong LIEN TIEP cung
       Vung thanh 1 diem (giu thoi diem som nhat), roi ap dung uu tien vung
       neo (MỎ* > BÃI RẠCH CHIẾC/BÃI TẬP KẾT) de tao dong Vung A -> Vung B.
       Gop tat ca xe, sap xep theo Thoi diem A tang dan.
    5. Ghi ban sao Excel local + upsert vao tab thang tren Google Sheets.

CAI DAT (chay 1 lan):
    pip install playwright pandas openpyxl gspread google-auth shapely
    playwright install chromium

CHAY THU CONG:
    python adsun_daily_report.py

DAT LICH CHAY TU DONG MOI GIO (Windows Task Scheduler) — xem huong dan hoi
lai de duoc ho tro tao/sua task.
"""

import argparse
import shutil
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright

import raw_cache
import sheets_client
import zone_matching
from adsun_common import (
    export_vehicle_report,
    get_credentials,
    goto_report_page,
    list_all_plates,
    login,
    make_scrape_lock_owner_id,
    wait_for_scrape_lock,
)

RAW_RANGE_PRESET = "3"  # "3 ngày" — khop voi MAX_LOOKBACK_HOURS=72h ben duoi, du du lieu
# tho de tu bu neu lan chay truoc gan nhat bi gian doan toi 3 ngay (xem giai
# thich o MAX_LOOKBACK_HOURS).
TIME_COLUMN = "Thời điểm"
ZONE_COLUMN = "Vùng"
HEADER_ROW_INDEX = 3  # dong 4 trong file excel (0-indexed) la dong ten cot

EMPTY_ZONE_VALUES = {"", "-", "–", "—"}

FINAL_COLUMNS = [
    "STT",
    "Thời điểm A",
    "Thời điểm B",
    "Biển số xe",
    "Vùng A",
    "Vùng B",
    "Tọa độ A",
    "Tọa độ B",
    "Địa chỉ A",
    "Địa chỉ B",
    "Nhiên liệu A",
    "Nhiên liệu B",
]

HIDDEN_COLUMNS = [
    "Tọa độ A",
    "Tọa độ B",
    "Địa chỉ A",
    "Địa chỉ B",
    "Nhiên liệu A",
    "Nhiên liệu B",
]

# Chi 6 cot nay duoc day len Google Sheets (spreadsheet bao cao rieng, xem
# sheets_client.get_report_spreadsheet) — Toa do/Dia chi/Nhien lieu van con
# day du trong ban Excel local (HIDDEN_COLUMNS o tren) de tham khao khi can.
REPORT_COLUMNS = [
    "STT",
    "Thời điểm A",
    "Thời điểm B",
    "Biển số xe",
    "Vùng A",
    "Vùng B",
]


ROLLING_WINDOW_HOURS = 24  # xem giai thich trong compute_window() ve ly do can rong toi 24h
# QUAN TRONG (16-17/8): may local NGU/TAT ~66 tieng (14/8 14:52 -> 17/8 09:07,
# xac nhan qua Windows Event Log — khong phai loi code) khien lich chay dinh
# ky local KHONG THUC THI duoc lan nao trong suot khoang do. Khi may tinh
# chay lai, cua so tu bu CU (24h) khong voi toi duoc phan du lieu qua xa
# (>24h) nen mat vinh vien cho toi khi phat hien va vao tay. RAW_RANGE_PRESET
# ("2" = 2 ngay ~ 48h) da san du lieu tho de tinh bu xa hon MA KHONG TON
# THEM CHI PHI QUET — chi can nang gioi han O DAU RA. Nang MAX_LOOKBACK_HOURS
# len 72h (rong hon ca 48h du lieu tho hien co) de neu may tinh gian doan
# vai ngay (ngu/tat/mat dien) roi chay lai, lan chay thanh cong dau tien co
# co hoi tu bu XA HON — khong con phai cho toi khi co nguoi bao lai.
MAX_LOOKBACK_HOURS = 72  # gioi han an toan khi tu dong mo rong cua so bu khoang trong


def compute_window(now: datetime, rolling_hours=ROLLING_WINDOW_HOURS, max_lookback_hours=MAX_LOOKBACK_HOURS):
    """Tra ve (window_start, now). Mac dinh window_start = now - rolling_hours
    ("N gio gan nhat", hien la 24 gio) — NHUNG se TU DONG MO RONG lui xa hon
    NEU lan chay dinh ky THANH CONG gan nhat (xem sheets_client.
    get_last_schedule_checkpoint) cach day LAU HON rolling_hours. Gioi han
    toi da max_lookback_hours (mac dinh 24 gio) de tranh mo rong vo han neu
    checkpoint bi hong/chua tung ghi (vd lan dau chay moi) — vuot qua muc do
    van can dung backfill_month.py/refresh_range.py thu cong nhu truoc.

    Vi sao rolling_hours mac dinh RONG toi 24h (khong chi vai gio): moi dong
    bao cao chi duoc tinh (va co the bi GHI DE lai) khi "Thời điểm A" (luc
    RỜI vung neo) nam trong window — nhung 1 chuyen di chi TRO THANH biet
    duoc (co Thời điểm B) khi xe DEN vung ke tiep, co the CACH RAT XA luc
    roi vung neo (vd xe roi bai luc 23h dem hom truoc, chay lung tung qua
    nhieu duong khong co vung ve, mai 4-5 tieng sau moi toi 1 vung co ve —
    da gap thuc te ngay 4/8, xe 51L97909 roi BÃI TẬP KẾT 23:23 hom truoc,
    toi MEKONG STAR PHÚ HỮU 04:38 hom sau, cach nhau hon 5 tieng). Neu
    rolling_hours qua hep (vd 3h), DEN LUC B xay ra thi A da "gia" hon ca
    do rong cua so, nen KHONG CO LAN CHAY DINH KY NAO (du chay dung gio,
    khong loi gi ca) co the bat duoc dong nay nua — khac voi truong hop
    checkpoint o tren (schedule bi tre/dut quang), day la gioi han cau truc
    xay ra CA KHI schedule chay hoan hao. Vi upsert_range_report doc/ghi
    LAI CA TAB moi lan (chi phi khong doi du window rong hay hep — xem
    docstring ham do), mo rong rolling_hours khong ton them chi phi API,
    nen dat bang luon max_lookback_hours cho an toan toi da."""
    default_start = now - timedelta(hours=rolling_hours)
    try:
        checkpoint = sheets_client.get_last_schedule_checkpoint()
    except Exception:
        checkpoint = None

    if checkpoint is not None and checkpoint < default_start:
        earliest_allowed = now - timedelta(hours=max_lookback_hours)
        window_start = max(checkpoint, earliest_allowed)
    else:
        window_start = default_start
    return window_start, now


def export_all_vehicles(page, raw_dir: Path, range_preset=RAW_RANGE_PRESET):
    plates = list_all_plates(page)
    if not plates:
        raise RuntimeError("Không lấy được danh sách xe từ dropdown 'Thiết bị'.")
    print(f"Tìm thấy {len(plates)} xe: {', '.join(plates)}")

    raw_dir.mkdir(parents=True, exist_ok=True)
    results = []  # list of (plate, path or None)
    seen_fingerprints = set()  # tich luy ca lan chay — xem adsun_common.export_vehicle_report

    for plate in plates:
        try:
            day_df, _fp = export_vehicle_report(
                page, plate, range_preset, HEADER_ROW_INDEX, seen_fingerprints
            )

            if day_df is None:
                print(f"  [{plate}] Không có dữ liệu, bỏ qua.")
                results.append((plate, None))
                continue

            raw_path = raw_dir / f"{plate}.xlsx"
            day_df.to_excel(raw_path, index=False)
            print(f"  [{plate}] Đã xuất tạm: {raw_path.name}")
            results.append((plate, raw_path))

            # Boi cache cho cac ngay DA QUA (khong bao gio cache "hom nay" vi
            # con thay doi) — tan dung du lieu vua quet ma khong ton them chi
            # phi, giup lan "lam moi vi them vung" sau nay nhanh hon nhieu.
            try:
                n = raw_cache.cache_complete_past_days(plate, day_df, date.today())
                if n:
                    print(f"  [{plate}] Đã bồi thêm {n} ngày vào cache.")
            except Exception as e:
                print(f"  [{plate}] Không bồi được cache: {e}", file=sys.stderr)

        except Exception as e:
            print(f"  [{plate}] LỖI khi xuất, bỏ qua xe này: {e}", file=sys.stderr)
            # Dung chuoi "ERROR" (khac None) de phan biet voi truong hop
            # "khong co du lieu that su" (has_no_data) — xe loi KHONG duoc
            # dua vao danh sach thay du lieu (xem only_replace_plates o main()),
            # tranh xoa mat du lieu cu con dang dung cua xe do.
            results.append((plate, "ERROR"))

    return results


def is_empty_zone(value):
    if pd.isna(value):
        return True
    return str(value).strip() in EMPTY_ZONE_VALUES


MAX_NOISE_EXCURSION_DEGREES = 0.003  # ~300m — xem extract_zone_points


def extract_zone_points(df, max_excursion_degrees=MAX_NOISE_EXCURSION_DEGREES):
    """Loc bo cac dong khong co Vung, roi gop cac dong cung mot Vung thanh 1
    diem duy nhat — nhung GIU LAI CA HAI moc thoi gian cua nhom do: thoi diem
    DEN (ping dau tien, luu o TIME_COLUMN nhu truoc) va thoi diem RỜI (ping
    CUOI CUNG con o vung do truoc khi doi vung, luu o "_leave"). build_
    transition_rows se dung thoi diem RỜI cho Vung A (gio lay = luc xe THUC
    SU roi khoi noi lay hang) va thoi diem DEN cho Vung B (gio giao).

    QUAN TRONG — phan biet 2 loai "mat vung" giua 2 lan cung 1 Vung:
    1. NHIEU GPS/canh vung (xe dung yen, dinh vi chi chap chon giua "trong
       vung" va "khong xac dinh" do sai so hoac vi tri sat ranh gioi ve) —
       PHAI GOP lai thanh 1 diem lien tuc, khong duoc tach.
    2. XE THUC SU ROI DI (lai sang noi khac) roi QUAY LAI dung vung do —
       PHAI tach thanh 2 diem rieng, moi lan ghe la 1 su kien khac nhau.
    Phan biet bang KHOANG CACH: neu MOI ping "khong xac dinh vung" trong
    khoang trong do van nam GAN vi tri cuoi cung ghi nhan trong vung (trong
    pham vi max_excursion_degrees, mac dinh ~300m), coi la nhieu GPS -> gop.
    Neu CO IT NHAT 1 ping di xa hon nguong do, coi la roi that su -> tach.
    (Chi dua vao "co dong nao xen giua khong" — khong xet khoang cach — se
    sai theo 2 huong: da gap thuc te xe 51M68032 ngay 7/8 ghe BÊ TÔNG ĐẠI
    THỊNH 19:39-19:48, LAI DI THAT hon 1km, quay lai 20:45-20:56 — can TACH;
    nhung cung ngay do, KÊNH A41 dau ngay 8/8 co 3 lan "vao/ra" vung chi vi
    dinh vi chap chon trong luc xe DUNG YEN tai 1 cho — can GOP, khong duoc
    tach thanh 3 diem gia.)"""
    points = []
    anchor_coord = None  # toa do CUOI CUNG ghi nhan luc CON TRONG 1 vung (KHONG doi trong luc "mat vung")
    max_gap_dist = 0.0  # khoang cach XA NHAT tinh TU anchor_coord ma cac ping "khong xac dinh vung" giua 2 lan da di toi
    for _, row in df.iterrows():
        zone = row[ZONE_COLUMN]
        coord = zone_matching.parse_coordinate(row.get("Tọa độ"))

        if is_empty_zone(zone):
            if coord is not None and anchor_coord is not None:
                dist = ((coord[0] - anchor_coord[0]) ** 2 + (coord[1] - anchor_coord[1]) ** 2) ** 0.5
                max_gap_dist = max(max_gap_dist, dist)
            continue  # KHONG cap nhat anchor_coord o day — phai giu co dinh xuyen suot khoang trong

        snapshot = {
            TIME_COLUMN: row[TIME_COLUMN],
            "Tọa độ": row.get("Tọa độ"),
            "Địa chỉ": row.get("Địa chỉ"),
            "Nhiên liệu": row.get("Nhiên liệu"),
        }
        same_zone_as_last_point = points and points[-1][ZONE_COLUMN] == zone
        is_noise_gap = max_gap_dist <= max_excursion_degrees
        if same_zone_as_last_point and is_noise_gap:
            points[-1]["_leave"] = snapshot  # nhieu GPS/canh vung -> van tinh la lien tuc
        else:
            points.append({**snapshot, ZONE_COLUMN: zone, "_leave": snapshot})
        if coord is not None:
            anchor_coord = coord  # cap nhat lai moc so sanh (vua xac nhan lai vi tri trong vung)
        max_gap_dist = 0.0
    return points


# Vung "neo" (bai/mo) mac dinh luon o vi tri Vung A: vung bat dau bang "MỎ",
# hoac dung "BÃI RẠCH CHIẾC" / "BÃI TẬP KẾT". Cac vung khac (Fico Mai Chi Tho,
# Fico Cat Lai...) mac dinh la Vung B.
ANCHOR_EXACT = {"BÃI RẠCH CHIẾC", "BÃI TẬP KẾT"}


def _name_upper(p):
    return str(p[ZONE_COLUMN]).strip().upper()


def is_anchor_default(name):
    name_upper = str(name).strip().upper()
    return name_upper.startswith("MỎ") or name_upper in ANCHOR_EXACT


def _forced_close_as_b(name_upper, prev_upper):
    """Ngoai le rieng cho BÃI TẬP KẾT: neu diem NGAY TRUOC no la 1 vung "MỎ*"
    (xe tu mo cho hang ve bai), thi lan nay BÃI TẬP KẾT dong vai Vung B (khep
    chang tu mo) thay vi vai Vung A mac dinh cua no. Neu sau do xe di tiep
    sang vung khac, BÃI TẬP KẾT lai dong vai Vung A (mac dinh) de mo chang moi."""
    return name_upper == "BÃI TẬP KẾT" and prev_upper is not None and prev_upper.startswith("MỎ")


def _pair_touches(points, idx, pair_rules_upper):
    """True neu diem points[idx] la 1 dau (A hoac B) cua 1 cap tuy chinh nguoi
    dung da khai bao, xet voi diem NGAY TRUOC hoac NGAY SAU no."""
    name_upper = _name_upper(points[idx])
    if idx > 0 and (_name_upper(points[idx - 1]), name_upper) in pair_rules_upper:
        return True
    if idx + 1 < len(points) and (name_upper, _name_upper(points[idx + 1])) in pair_rules_upper:
        return True
    return False


def _pick_group_representative(points, i, j, easy_pass_zones):
    """Chon diem dai dien cho 1 nhom vai tro lien tiep points[i..j]: uu tien
    diem KHONG thuoc easy_pass_zones ("vung de di qua", danh dau tren tab Ve
    vung) neu nhom co it nhat 1 diem nhu vay — vd nhom [FICO CÁT LÁI,
    YAMAKEN CÁT LÁI, FICO CÁT LÁI] voi FICO CÁT LÁI duoc danh dau "de di qua"
    se chon YAMAKEN CÁT LÁI lam diem dai dien, vi FICO chi la diem BUOC PHAI
    di ngang qua de toi/roi khoi YAMAKEN roi quay ve. Neu CA NHOM chi toan
    vung de-di-qua (vd xe chi ghe DUY NHAT vung do roi quay ve, khong ghe
    vung nao khac xen giua), van chon diem CUOI CUNG nhu binh thuong — luc
    nay chinh no moi la diem dung that su."""
    non_pass = [k for k in range(i, j + 1) if _name_upper(points[k]) not in easy_pass_zones]
    candidates = non_pass if non_pass else list(range(i, j + 1))
    return points[candidates[-1]]


def _collapse_runs(points, pair_rules_upper, easy_pass_zones):
    """Gop cac diem LIEN TIEP cung vai tro mac dinh (A voi A, B voi B) thanh 1
    diem duy nhat (xem _pick_group_representative de biet cach chon diem dai
    dien trong nhom) — vi du xe di ngang qua nhieu vung dich lien tiep truoc
    khi quay lai dung vung neo xuat phat, chi vung dich THAT SU (khong phai
    vung "de di qua") moi la diem dung that su. Cac diem la "diem noi" (hinge
    — hoac bi ep dong vai B nhu BÃI TẬP KẾT sau MỎ, hoac la 1 dau cua cap tuy
    chinh) khong bao gio bi gop, luon dung rieng le."""
    reduced = []
    n = len(points)
    i = 0
    while i < n:
        name_upper = _name_upper(points[i])
        prev_upper = _name_upper(points[i - 1]) if i > 0 else None
        forced = _forced_close_as_b(name_upper, prev_upper)
        boundary = forced or _pair_touches(points, i, pair_rules_upper)

        if boundary:
            reduced.append(points[i])
            i += 1
            continue

        role = "A" if is_anchor_default(points[i][ZONE_COLUMN]) else "B"
        j = i
        while j + 1 < n:
            nxt_upper = _name_upper(points[j + 1])
            nxt_prev_upper = _name_upper(points[j])
            nxt_forced = _forced_close_as_b(nxt_upper, nxt_prev_upper)
            nxt_boundary = nxt_forced or _pair_touches(points, j + 1, pair_rules_upper)
            nxt_role = "A" if is_anchor_default(points[j + 1][ZONE_COLUMN]) else "B"
            if nxt_boundary or nxt_role != role:
                break
            j += 1
        reduced.append(_pick_group_representative(points, i, j, easy_pass_zones))
        i = j + 1
    return reduced


def build_transition_rows(points, plate, pair_rules=None, easy_pass_zones=None):
    """Tao cac dong chuyen vung A->B tu day diem vung da loc (xem
    _collapse_runs de biet buoc gop truoc). Duyet day da gop, giu 1 "diem dang
    mo" (ung vien Vung A):
      - Neu (diem dang mo -> diem hien tai) khop 1 cap tuy chinh nguoi dung
        khai bao -> LUON tao dong, bat ke vai tro mac dinh.
      - Neu diem hien tai la vung neo mac dinh (va khong bi ep dong vai B) ->
        cap nhat lam diem dang mo moi (gop chuoi neo lien tiep).
      - Nguoc lai (dong vai Vung B): tao dong (diem dang mo -> diem hien tai)
        neu da co diem dang mo; sau do diem nay tro thanh diem dang mo MOI chi
        khi no se "mo lai" duoc (bi ep dong vai B nhu BÃI TẬP KẾT, hoac la dau
        A cua 1 cap tuy chinh voi diem ke tiep) — neu khong thi dong lai (ve
        vung neo, khong ghi)."""
    pair_rules_upper = {(a.strip().upper(), b.strip().upper()) for a, b in (pair_rules or [])}
    reduced = _collapse_runs(points, pair_rules_upper, easy_pass_zones or set())

    rows = []
    open_point = None
    n = len(reduced)

    def make_row(a, b):
        a_leave = a["_leave"]  # gio lay = luc RỜI khoi Vung A (ping cuoi cung con o do)
        return {
            "Thời điểm A": a_leave[TIME_COLUMN],
            "Thời điểm B": b[TIME_COLUMN],
            "Biển số xe": plate,
            "Vùng A": a[ZONE_COLUMN],
            "Vùng B": b[ZONE_COLUMN],
            "Tọa độ A": a_leave["Tọa độ"],
            "Tọa độ B": b["Tọa độ"],
            "Địa chỉ A": a_leave["Địa chỉ"],
            "Địa chỉ B": b["Địa chỉ"],
            "Nhiên liệu A": a_leave["Nhiên liệu"],
            "Nhiên liệu B": b["Nhiên liệu"],
        }

    for idx, p in enumerate(reduced):
        name_upper = _name_upper(p)
        prev_upper = _name_upper(reduced[idx - 1]) if idx > 0 else None
        nxt_upper = _name_upper(reduced[idx + 1]) if idx + 1 < n else None
        forced_close_as_b = _forced_close_as_b(name_upper, prev_upper)

        pair_override = open_point is not None and (_name_upper(open_point), name_upper) in pair_rules_upper
        if pair_override:
            rows.append(make_row(open_point, p))
            will_reopen = (
                forced_close_as_b
                or (nxt_upper is not None and (name_upper, nxt_upper) in pair_rules_upper)
                or (is_anchor_default(p[ZONE_COLUMN]) and not forced_close_as_b)
            )
            open_point = p if will_reopen else None
            continue

        if is_anchor_default(p[ZONE_COLUMN]) and not forced_close_as_b:
            open_point = p
            continue

        if open_point is None:
            continue  # chua biet xe xuat phat tu vung neo nao, bo qua

        rows.append(make_row(open_point, p))
        will_reopen = forced_close_as_b or (nxt_upper is not None and (name_upper, nxt_upper) in pair_rules_upper)
        open_point = p if will_reopen else None

    return rows


def process_vehicle_dataframe(
    plate, df, window_start, window_end, zone_polygons=None, pair_rules=None, easy_pass_zones=None
):
    """Tinh Vung + tao cac dong chuyen vung cho 1 xe TU TOAN BO du lieu ping
    truyen vao (KHONG loc theo khung gio truoc khi tinh), roi CHI LOC CAC
    DONG KET QUA (transition) theo "Thời điểm A" nam trong [window_start,
    window_end). Dung chung cho ca luong doc tu file tam (adsun_daily_report)
    va luong doc tu cache/DataFrame trong bo nho (backfill_month, refresh_range).

    QUAN TRONG: truoc day ham nay loc RAW PING truoc (df = df[window_start <=
    t < window_end]) roi moi tinh diem/chuyen vung — dieu nay lam GAY mat 1
    chuyen di neu diem KET THUC (Thời điểm B) roi ra ngoai window nhung diem
    BAT DAU (Thời điểm A) van nam trong window (vi ping ket thuc bi cat mat,
    "diem dang mo" khong bao gio duoc dong lai nen khong sinh ra dong nao).
    Day chinh la nguyen nhan lam mat dong 51M65797 09:57:26 -> 10:28:01 khi
    chay refresh_range.py voi --to "...10:00" (dong bi cat vi 10:28:01 > 10:00).
    Sua bang cach: tinh transition tu FULL du lieu ping duoc truyen vao (caller
    chiu trach nhiem truyen du du lieu, vd ca ngay/nhieu ngay), chi loc O DAU
    RA theo Thời điểm A — vay du chuyen di bat dau trong window luon duoc tinh
    dung du no ket thuc sau window bao nhieu."""
    if TIME_COLUMN not in df.columns or ZONE_COLUMN not in df.columns:
        print(f"  [{plate}] Thiếu cột cần thiết, bỏ qua.", file=sys.stderr)
        return []

    if df.empty:
        return []
    df = df.sort_values(TIME_COLUMN)

    if zone_polygons is not None:
        # Ghi de cot Vung cua Adsun bang vung tu tinh tu toa do — giai
        # quyet van de Adsun khong hoi to vung moi tao cho du lieu qua khu.
        df = zone_matching.compute_zone_column(df, zone_polygons)

    points = extract_zone_points(df)
    rows = build_transition_rows(points, plate, pair_rules, easy_pass_zones)
    return [r for r in rows if window_start <= r["Thời điểm A"] < window_end]


def merge_reports(results, window_start, window_end, zone_polygons=None, pair_rules=None, easy_pass_zones=None):
    all_rows = []
    for plate, raw_path in results:
        if raw_path is None or raw_path == "ERROR":
            continue
        try:
            # raw_path duoc ghi lai boi export_all_vehicles TU DataFrame da
            # doc san (header=HEADER_ROW_INDEX ap dung luc do roi) — o day chi
            # can doc lai voi header mac dinh (dong dau tien).
            df = pd.read_excel(raw_path)
        except Exception as e:
            print(f"  [{plate}] LỖI khi đọc file tạm {raw_path.name}: {e}", file=sys.stderr)
            continue

        all_rows.extend(
            process_vehicle_dataframe(plate, df, window_start, window_end, zone_polygons, pair_rules, easy_pass_zones)
        )

    if not all_rows:
        return None

    merged = pd.DataFrame(all_rows)
    merged = merged.sort_values("Thời điểm A").reset_index(drop=True)
    merged.insert(0, "STT", range(1, len(merged) + 1))
    return merged[FINAL_COLUMNS]


def format_workbook(path, hidden_column_names):
    """An cac cot chi dinh, va chinh do rong tat ca cot khop voi do dai ky tu
    dai nhat trong cot do (ke ca header)."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    header = [c.value for c in ws[1]]

    for name in hidden_column_names:
        if name in header:
            col_letter = get_column_letter(header.index(name) + 1)
            ws.column_dimensions[col_letter].hidden = True

    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(path)


def upsert_split_by_month(merged, range_start, range_end, only_replace_plates=None):
    """upsert_range_report ghi theo 1 tab-thang duy nhat — neu khoang can cap
    nhat vo tinh vat qua ranh gioi thang (hiem, chi xay quanh nua dem dau
    thang do dung cua so lan can "N gio gan nhat"), chia nho ra ghi dung
    tung tab thang tuong ung, tranh ghi nham/mat du lieu thang ke tiep.

    only_replace_plates: xem docstring cua sheets_client.upsert_range_report
    — BAT BUOC truyen dung tap bien so DA QUET THANH CONG lan nay, neu khong
    nhung xe bi loi/bo qua se bi XOA MAT du lieu cu thay vi duoc giu nguyen."""
    boundaries = [range_start]
    cursor = range_start.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    while cursor <= range_end:
        next_month = cursor.month + 1 if cursor.month < 12 else 1
        next_year = cursor.year if cursor.month < 12 else cursor.year + 1
        cursor = cursor.replace(year=next_year, month=next_month)
        if range_start < cursor < range_end:
            boundaries.append(cursor)
    boundaries.append(range_end)
    boundaries = sorted(set(boundaries))

    for i in range(len(boundaries) - 1):
        seg_start, seg_end = boundaries[i], boundaries[i + 1]
        seg_rows = merged[(merged["Thời điểm A"] >= seg_start) & (merged["Thời điểm A"] < seg_end)]
        tab_name = f"{seg_start:%Y-%m}"
        sheets_client.upsert_range_report(
            tab_name, seg_start, seg_end, seg_rows[REPORT_COLUMNS], only_replace_plates=only_replace_plates
        )
        print(
            f"Đã cập nhật tab '{tab_name}' cho khoảng {seg_start:%d/%m/%Y %H:%M} -> "
            f"{seg_end:%d/%m/%Y %H:%M} ({len(seg_rows)} dòng)."
        )


def main():
    parser = argparse.ArgumentParser(
        description="Xuất báo cáo hành trình tổng hợp toàn bộ xe (8h hôm qua - 8h hôm nay)"
    )
    parser.add_argument(
        "--output-dir",
        default=str(Path(__file__).parent / "reports"),
        help="Thư mục lưu file Excel tổng hợp (mặc định: ./reports)",
    )
    parser.add_argument(
        "--show", action="store_true", help="Hiện trình duyệt (mặc định chạy ẩn, phù hợp chạy tự động)"
    )
    parser.add_argument(
        "--keep-raw", action="store_true", help="Giữ lại các file Excel riêng từng xe (để kiểm tra)"
    )
    args = parser.parse_args()

    now = datetime.now()
    window_start, window_end = compute_window(now)
    print(f"Khung thời gian báo cáo ({ROLLING_WINDOW_HOURS} giờ gần nhất): {window_start:%d/%m/%Y %H:%M} -> {window_end:%d/%m/%Y %H:%M}")

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    raw_dir = output_dir / "_raw" / f"{window_start:%Y%m%d}"

    username, password = get_credentials()

    # Khoa quet Adsun toan cuc (chung voi may local + GitHub Actions) — tranh
    # 2 phien dang nhap Adsun DONG THOI gay nham lan du lieu giua cac xe (da
    # gap thuc te nhieu lan, xem adsun_common.acquire_scrape_lock).
    lock_owner = make_scrape_lock_owner_id()
    if not wait_for_scrape_lock(lock_owner):
        print(
            "Đang có tiến trình khác quét Adsun, bỏ qua lần chạy này (lần chạy kế "
            "tiếp sẽ tự bù nhờ cửa sổ báo cáo rộng).",
            file=sys.stderr,
        )
        return

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=not args.show)
            context = browser.new_context(accept_downloads=True)
            page = context.new_page()

            try:
                login(page, username, password)
                print("Đăng nhập thành công.")

                goto_report_page(page)
                results = export_all_vehicles(page, raw_dir)

            except Exception as e:
                page.screenshot(path="error.png")
                print(f"\nLỗi: {e}\nĐã lưu ảnh chụp màn hình error.png để kiểm tra.", file=sys.stderr)
                sys.exit(1)
            finally:
                browser.close()
    finally:
        try:
            sheets_client.release_scrape_lock(lock_owner)
        except Exception:
            pass

    # Danh dau "da kiem tra xong toi window_end" NGAY KHI buoc quet thanh
    # cong (bat ke sau do co du lieu hay khong) — dung lam moc de lan chay
    # ke tiep tu dong phat hien va bu khoang trong neu gap su co (xem
    # compute_window). Ghi CANG SOM CANG TOT (truoc khi xu ly Vung/gop du
    # lieu) de dam bao checkpoint luon phan anh dung "da quet xong", khong
    # bi anh huong boi loi o cac buoc xu ly phia sau.
    try:
        sheets_client.set_last_schedule_checkpoint(window_end)
    except Exception as e:
        print(f"Không ghi được checkpoint lịch chạy ({e}).", file=sys.stderr)

    try:
        zone_polygons = zone_matching.load_zone_polygons()
        print(f"Đã tải {len(zone_polygons)} vùng từ Google Sheets để tự tính Vùng theo tọa độ.")
    except Exception as e:
        zone_polygons = None
        print(
            f"Không tải được danh sách vùng từ Google Sheets ({e}). "
            "Dùng tạm cột Vùng có sẵn của Adsun.",
            file=sys.stderr,
        )

    try:
        pair_rules = sheets_client.list_pair_rules()
        print(f"Đã tải {len(pair_rules)} cặp vùng tùy chỉnh.")
    except Exception as e:
        pair_rules = []
        print(f"Không tải được danh sách cặp vùng tùy chỉnh ({e}).", file=sys.stderr)

    try:
        easy_pass_zones = zone_matching.load_easy_pass_zones()
        if easy_pass_zones:
            print(f"Đã tải {len(easy_pass_zones)} vùng \"dễ đi qua\".")
    except Exception as e:
        easy_pass_zones = set()
        print(f"Không tải được danh sách vùng \"dễ đi qua\" ({e}).", file=sys.stderr)

    # Chi nhung xe QUET THANH CONG lan nay (co du lieu hoac xac nhan khong co
    # du lieu that su) moi duoc phep thay du lieu cu — xe bi loi ("ERROR")
    # se duoc GIU NGUYEN du lieu cu, khong bi xoa mat (xem upsert_split_by_month).
    successful_plates = {plate for plate, status in results if status != "ERROR"}

    print("Đang gộp và lọc dữ liệu theo đúng khung giờ...")
    merged = merge_reports(results, window_start, window_end, zone_polygons, pair_rules, easy_pass_zones)

    if merged is None or merged.empty:
        print("Không có dữ liệu hành trình nào trong khung thời gian này cho toàn bộ xe.")
        return

    output_path = output_dir / f"BaoCaoHanhTrinhTongHop_{window_start:%Y-%m-%d}.xlsx"
    try:
        merged.to_excel(output_path, index=False)
    except PermissionError:
        # File dang bi mo (vd. dang mo trong Excel) nen khong ghi de duoc,
        # luu sang ten khac de khong mat du lieu ngay hom nay.
        output_path = output_dir / f"BaoCaoHanhTrinhTongHop_{window_start:%Y-%m-%d}_{datetime.now():%H%M%S}.xlsx"
        merged.to_excel(output_path, index=False)
        print(
            f"File gốc đang bị mở (có thể đang mở trong Excel), đã lưu vào tên khác: {output_path.name}",
            file=sys.stderr,
        )
    format_workbook(output_path, HIDDEN_COLUMNS)
    print(f"Đã lưu báo cáo tổng hợp: {output_path.resolve()} ({len(merged)} dòng)")

    try:
        # Dung window_end (= "now" thuc su, tra ve tu compute_window), KHONG
        # phai window_start + 1 ngay — nham lan nay se lam upsert ghi de mot
        # khoang thoi gian tuong lai rat lon (loi da tung xay ra khi doi sang
        # cua so "N gio gan nhat" ma quen sua cho nay).
        upsert_split_by_month(merged[FINAL_COLUMNS], window_start, window_end, only_replace_plates=successful_plates)
    except Exception as e:
        print(f"Không đẩy được báo cáo lên Google Sheets: {e}", file=sys.stderr)

    try:
        # Danh dau "done" cho co che nut "Cap nhat ngay" tren web app (neu co
        # ai vua yeu cau) — ghi lai du la lan chay nay tu dong theo lich hay
        # duoc kich hoat thu cong, khong anh huong gi neu khong ai yeu cau.
        sheets_client.set_refresh_status("done", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    except Exception:
        pass

    if not args.keep_raw:
        shutil.rmtree(raw_dir, ignore_errors=True)


if __name__ == "__main__":
    main()
